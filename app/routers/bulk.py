"""Bulk product upload via Excel file.
Flow: receives Excel file → stores to Minio → reads back → parses → validates → bulk INSERT
"""

import io
import logging
from datetime import datetime

import openpyxl
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from minio import Minio
from psycopg2.extras import Json

from ..config import settings
from ..db import get_db
from ..auth import verify_token

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/products", tags=["Bulk Upload"])

def get_minio() -> Minio:
    return Minio(
        settings.MINIO_ENDPOINT,
        access_key=settings.MINIO_ACCESS_KEY,
        secret_key=settings.MINIO_SECRET_KEY,
        secure=False,
    )

TEMPLATE_COLUMNS = ["SKU", "Name", "Unit", "Price", "Category", "Attributes"]

@router.get("/template")
def download_template(user: dict = Depends(verify_token)):
    """Download Excel template for bulk product upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(TEMPLATE_COLUMNS)
    ws.append(["BRG-001", "Beras 5kg", "pcs", 65000, "Sembako", '{"berat":"5kg"}'])
    ws.append(["BRG-002", "Gula 1kg", "pcs", 14000, "Sembako", "{}"])

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=product-template.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )

@router.post("/bulk-upload")
async def bulk_upload(file: UploadFile = File(...), user: dict = Depends(verify_token)):
    """Receive Excel file, store to Minio, then process for bulk product insert."""
    if not (file.filename and (file.filename.endswith(".xlsx") or file.filename.endswith(".xls") or file.filename.endswith(".csv"))):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted")

    content = await file.read()
    file_size = len(content)
    if file_size == 0:
        raise HTTPException(status_code=400, detail="Empty file")
    if file_size > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    object_name = f"bulk/{timestamp}_{file.filename}"

    try:
        mc = get_minio()
        mc.put_object(
            settings.MINIO_BUCKET,
            object_name,
            io.BytesIO(content),
            length=file_size,
            content_type=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        logger.info("Stored %s to Minio", object_name)
    except Exception as e:
        logger.error("Minio upload failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Failed to store file: {e}")

    try:
        mc = get_minio()
        response = mc.get_object(settings.MINIO_BUCKET, object_name)
        file_data = response.read()
        response.close()
        response.release_conn()
    except Exception as e:
        logger.error("Minio read failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Failed to read file from storage: {e}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        logger.error("Excel parse failed: %s", e)
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="No data rows found (header only)")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    logger.info("Excel header: %s", header)

    col_map = _map_columns(header)
    if col_map is None:
        raise HTTPException(status_code=400, detail="Missing required columns: SKU, Name. Found: " + ", ".join(header))

    errors = []
    products = []
    for i, row in enumerate(rows[1:], start=2):
        row_errors = _validate_row(row, col_map, i)
        if row_errors:
            errors.append({"row": i, "sku": _get_sku(row, col_map), "reasons": row_errors})
        else:
            products.append(_build_product(row, col_map))

    results = []
    conn = get_db()
    cur = conn.cursor()
    try:
        for prod in products:
            try:
                cur.execute(
                    """INSERT INTO products (sku, name, unit, price, attributes)
                       VALUES (%s, %s, %s, %s, %s)
                       RETURNING id, sku, name""",
                    (prod["sku"], prod["name"], prod["unit"], prod.get("price", 0.0), Json(prod["attributes"])),
                )
                row = cur.fetchone()
                conn.commit()
                results.append({"row": prod["_row"], "sku": prod["sku"], "name": prod["name"], "status": "imported"})
            except Exception as e:
                conn.rollback()
                err_msg = str(e)
                if "duplicate key" in err_msg.lower() or "unique" in err_msg.lower():
                    results.append({"row": prod["_row"], "sku": prod["sku"], "name": prod["name"], "status": "skipped", "reason": "SKU already exists"})
                else:
                    results.append({"row": prod["_row"], "sku": prod["sku"], "name": prod["name"], "status": "error", "reason": err_msg})
    finally:
        conn.close()

    for err in errors:
        results.append({
            "row": err["row"],
            "sku": err["sku"],
            "name": "",
            "status": "error",
            "reason": "; ".join(err["reasons"]),
        })

    results.sort(key=lambda r: r["row"])

    return {
        "status": "ok",
        "total": len(results),
        "imported": sum(1 for r in results if r["status"] == "imported"),
        "skipped": sum(1 for r in results if r["status"] == "skipped"),
        "errors": sum(1 for r in results if r["status"] == "error"),
        "file": object_name,
        "results": results,
    }

def _map_columns(header: list[str]):
    mapping = {}
    for i, col in enumerate(header):
        if col in ("sku", "kode", "code"):
            mapping["sku"] = i
        elif col in ("name", "nama", "product", "produk"):
            mapping["name"] = i
        elif col in ("unit", "satuan", "uom"):
            mapping["unit"] = i
        elif col in ("price", "harga"):
            mapping["price"] = i
        elif col in ("category", "kategori", "cat"):
            mapping["category"] = i
        elif col in ("attributes", "atribut", "attr"):
            mapping["attributes"] = i

    if "sku" not in mapping or "name" not in mapping:
        return None
    return mapping

def _validate_row(row: tuple, col_map: dict, row_num: int) -> list[str]:
    errors = []
    sku_idx = col_map.get("sku")
    name_idx = col_map.get("name")

    if sku_idx is not None:
        sku = str(row[sku_idx]).strip() if row[sku_idx] is not None else ""
        if not sku:
            errors.append("SKU is required")
    if name_idx is not None:
        name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
        if not name:
            errors.append("Name is required")
    price_idx = col_map.get("price")
    if price_idx is not None and price_idx < len(row):
        val = row[price_idx]
        if val is not None:
            try:
                float(val)
            except (ValueError, TypeError):
                errors.append(f"Invalid price: {val}")
    return errors

def _get_sku(row: tuple, col_map: dict) -> str:
    idx = col_map.get("sku")
    if idx is not None and idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ""

def _build_product(row: tuple, col_map: dict) -> dict:
    prod = {"_row": 0, "sku": "", "name": "", "unit": "pcs", "attributes": {}, "category": ""}

    for key, idx in col_map.items():
        if idx >= len(row):
            continue
        val = row[idx]

        if key == "sku":
            prod["sku"] = str(val).strip() if val else ""
        elif key == "name":
            prod["name"] = str(val).strip() if val else ""
        elif key == "unit":
            prod["unit"] = str(val).strip() if val else "pcs"
        elif key == "price":
            try:
                prod["price"] = float(val) if val else 0.0
            except (ValueError, TypeError):
                prod["price"] = 0.0
        elif key == "attributes":
            if val:
                try:
                    import json
                    prod["attributes"] = json.loads(val) if isinstance(val, str) else val
                except (json.JSONDecodeError, TypeError):
                    prod["attributes"] = {}
        elif key == "category":
            prod["category"] = str(val).strip() if val else ""

    return prod